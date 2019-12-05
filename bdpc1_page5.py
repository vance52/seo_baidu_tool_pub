# ‐*‐ coding: utf‐8 ‐*‐
"""
监控某域名前五页排名变化，可实现分关键词类别监控
bdpc1_page5_info.txt记录每个kwd在第几页有排名
bdpc1_page5_rankurl.txt记录每个kwd前五页排名的url，当前页面出现多个就记录多个
生成的excel和txt是首页 前二/三/四/五/五页后的数量
bdpc1_page5_info.txt的行数就是真正查询成功的词数
excel文件kwd_core_city.xlsx 中每个sheet名代表一类关键词,每个sheet第一列放关键词
现在百度PC反爬有点严重,head头要加cookie,查询频繁还要轮换cookie
"""
import requests
from pyquery import PyQuery as pq
import threading
import queue
import time
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import gc


# 计算最终结果
def get_result(file_path, result, result_last):
    # 读取文件 计算首页到五页每页词数
    for line in open(file_path, 'r', encoding='utf-8'):
        line = line.strip().split('\t')
        page = line[1]
        group = line[2]
        result[group][page] += 1

    # 计算首页、前二页、前三页、前四页、前五页词数
    for group, page_num in result.items():
        print(group, page_num)
        for page, num in page_num.items():
            if page == '五页后':
                result_last[group]['五页后'] = num
            if page in ['首页', '二页', '三页', '四页', '五页']:
                result_last[group]['前五页'] += num
            if page in ['首页', '二页', '三页', '四页']:
                result_last[group]['前四页'] += num
            if page in ['首页', '二页', '三页']:
                result_last[group]['前三页'] += num
            if page in ['首页', '二页']:
                result_last[group]['前二页'] += num
            if page == '首页':
                result_last[group]['首页'] += num
    return result_last


# 写入excel文件
def write_excel(group_list, result_last, today):
    wb = Workbook()
    # 创建sheet
    for group in group_list:
        sheet_num = 0
        wb.create_sheet(u'{0}'.format(group), index=sheet_num)
        sheet_num += 1
    # 写内容
    row_first = ['日期', '首页', '前二页', '前三页', '前四页', '前五页', '五页后']
    for group, data_dict in result_last.items():
        # 写第一行表头
        wb[group].append(row_first)
        # 写数据
        row_value = [today]
        for page, value in data_dict.items():
            row_value.append(value)
        wb[u'{0}'.format(group)].append(row_value)
    wb.save('{0}bdpc1_page5.xlsx'.format(today))


class bdpcCoverPage5(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    @staticmethod
    def read_excel(filepath):
        q = queue.Queue()
        group_list = []
        kwd_dict = {}
        wb_kwd = load_workbook(filepath)
        for sheet_obj in wb_kwd:
            sheet_name = sheet_obj.title
            group_list.append(sheet_name)
            kwd_dict[sheet_name]= []
            col_a = sheet_obj['A']
            for cell in col_a:
                kwd = (cell.value)
                # 加个判断吧
                if kwd:
                    kwd_z = kwd+ '租房'
                    kwd_er = kwd + '二手房'
                    q.put([sheet_name,kwd_z])
                    q.put([sheet_name,kwd_er])
        return q, group_list

    # 初始化结果字典
    @staticmethod
    def result_init(group_list):
        pages = ['首页','二页','三页','四页','五页','五页后']
        result = {} #记录首页、第二页..第五页各多少词
        for group in group_list:
            result[group] = {}
        for group in group_list:
            for page in pages:
                result[group][page] = 0
        print("结果字典init...")
        return result

    # 初始化最后结果字典
    @staticmethod
    def result_last_init(group_list):
        result_last = {}  # 记录首页、第二页..第五页各多少词
        pages = ['首页','前二页','前三页','前四页','前五页','五页后']
        for group in group_list:
            result_last[group] = {}
        for group in group_list:
            for page in pages:
                result_last[group][page] = 0
        print("最终结果字典init...")
        return result_last

    # 获取某词serp源码
    def get_html(self,url,retry=2):
        try:
            r = requests.get(url=url,headers=user_agent,timeout=5)
        except Exception as e:
            print('获取源码失败',e)
            time.sleep(6)
            if retry > 0:
                self.get_html(url,retry-1)
        else:
            html = r.content.decode('utf-8',errors='ignore')  # 用r.text有时候识别错误
            url = r.url
            return html,url

    # 获取某词serp源码上自然排名的所有url
    def get_encrpt_urls(self,html,url):
        encrypt_url_list = []
        doc = pq(html)
        title = doc('title').text()
        if '_百度搜索' in title and 'https://www.baidu.com/s?tn=48020221' in url:
            try:
                a_list = doc('.t a').items()
            except Exception as e:
                print('未提取到serp上的解密url', e)
            else:
                for a in a_list:
                    encrypt_url = a.attr('href')
                    if encrypt_url.find('http://www.baidu.com/link?url=') == 0:
                        encrypt_url_list.append(encrypt_url)
        else:
            print(title,'源码异常,可能反爬')
            time.sleep(6)
        return encrypt_url_list

    # 解密某条加密url
    def decrypt_url(self,encrypt_url,retry=1):
        real_url = None # 默认None
        try:
            encrypt_url = encrypt_url.replace('http://','https://')
            # print(encrypt_url)
            r = requests.head(encrypt_url,headers=user_agent)
        except Exception as e:
            print(encrypt_url,'解密失败',e)
            time.sleep(6)
            if retry > 0:
                self.decrypt_url(encrypt_url,retry-1)
        else:
            real_url = r.headers['Location']
            return real_url

    # 获取某词serp源码首页排名真实url
    def get_real_urls(self,encrypt_url_list):
        real_url_list = [self.decrypt_url(encrypt_url) for encrypt_url in encrypt_url_list]
        real_url_set = set(real_url_list)
        real_url_set = real_url_set.remove(None) if None in real_url_set else real_url_set
        real_url_list = list(real_url_set)
        return real_url_list

    # 提取某url的域名部分
    def get_domain(self,real_url):
        domain = None
        try:
           res = urlparse(real_url)
        except Exception as e:
           print (e,real_url)
        else:
           domain = res.netloc
        return domain

    # 获取某词serp源码首页排名所有域名
    def get_domains(self,real_url_list):
            domain_list = [self.get_domain(real_url) for real_url in real_url_list]
            # 一个词某域名多个url有排名,计算一次
            domain_set = set(domain_list)
            domain_set = domain_set.remove(None) if None in domain_set else domain_set
            domain_str = ','.join(domain_set)
            return domain_str

    # 线程函数
    def run(self):
        while 1:
            group_kwd = q.get()
            group,kwd = group_kwd
            try:
                for page in page_dict.keys():
                    if page == '':
                        url = "https://www.baidu.com/s?tn=48020221_28_hao_pg&ie=utf-8&wd={0}".format(kwd)
                    else:
                        url = "https://www.baidu.com/s?tn=48020221_28_hao_pg&ie=utf-8&wd={0}&pn={1}".format(kwd,page)
                    # print(url)
                    html,now_url = self.get_html(url)
                    encrypt_url_list = self.get_encrpt_urls(html,now_url)
                    real_url_list = self.get_real_urls(encrypt_url_list)
                    # print(real_url_list)
                    domain_str = self.get_domains(real_url_list)
                    if domain_str and domain in domain_str:
                        page = '首页' if page == '' else page_dict[page]
                        f.write('{0}\t{1}\t{2}\n'.format(kwd,page,group))
                        for real_url in real_url_list:
                            if domain in real_url:
                                f_url.write('{0}\t{1}\t{2}\t{3}\n'.format(kwd,page,group,real_url))
                        break
                    if page == 40 and domain not in domain_str:
                        f.write('{0}\t{1}\t{2}\n'.format(kwd,'五页后',group))
                f.flush()
                f_url.flush()
            except Exception as e:
                print(e)
            finally:
                del kwd
                gc.collect()
                q.task_done()


if __name__ == "__main__":
    start = time.time()
    local_time = time.localtime()
    today = time.strftime('%Y%m%d',local_time)
    domain = '5i5j.com'
    user_agent = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
        'Host': 'www.baidu.com',
        'Referer': 'https://www.hao123.com/?tn=48020221_28_hao_pg',
        'Cookie':'BDUSS=RQSWpXeEd5a1lDLXF6OExrVmdzeW9WUnFqRWtTNGRpWEctOU1RSmw0bHB5WU5aSVFBQUFBJCQAAAAAAAAAAAEAAADlnXKtt7CwybK-vbYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGk8XFlpPFxZeG'
    }
    q,group_list = bdpcCoverPage5.read_excel('kwd_core_city.xlsx')  # 关键词队列及分类
    result = bdpcCoverPage5.result_init(group_list)  # 结果字典-记录首页到5页各多少词
    result_last = bdpcCoverPage5.result_last_init(group_list)  # 结果字典-统计首页,前2页到前5页多少词
    all_num = q.qsize() # 总词数
    page_dict = {'':'首页',10:'二页',20:'三页',30:'四页',40:'五页'}  # 查询页码
    f = open('{0}bdpc1_page5_info.txt'.format(today),'w',encoding="utf-8")
    f_url = open('{0}bdpc1_page5_rankurl.txt'.format(today),'w',encoding="utf-8")
    file_path = f.name
    # 设置线程数
    for i in list(range(1)):
        t = bdpcCoverPage5()
        t.setDaemon(True)
        t.start()
    q.join()
    f.close()
    f_url.close()
    # 根据bdpc1_page5_info.txt数据计算结果
    result_last = get_result(file_path,result,result_last)
    # 写入txt文件
    with open('{0}bdpc1_page5.txt'.format(today), 'w', encoding="utf-8") as f:
        for group,data_dict in result_last.items():
            for page,value in data_dict.items():
                f.write(group + '\t' + page + '\t' + str(value) + '\n')
    end = time.time()
    # 写入excel
    write_excel(group_list,result_last,today)
    print('关键词共{0}个,成功个数看page5_info.txt,耗时{1}min'.format(all_num, (end - start) / 60))
